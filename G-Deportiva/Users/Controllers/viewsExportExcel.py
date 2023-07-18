from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import io
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from rest_framework.decorators import api_view
from ..models import *
from ..serializers import *



@api_view(['GET'])
def export_users_to_excel(request):
   # Obtener los parámetros de la URL para filtrar los datos
    id = request.GET.get('id')
    email = request.GET.get('email')
    name = request.GET.get('name')
    last_name = request.GET.get('last_name')
    cod_rol = request.GET.get('cod_rol')
    last_name = request.GET.get('last_name')
    last_name = request.GET.get('last_name')
    # Agrega aquí los demás campos que desees filtrar

    # Filtrar los datos en función de los parámetros
    queryset = User.objects.all()
    if email:
        queryset = queryset.filter(email=email)
    if name:
        queryset = queryset.filter(name=name)
    if last_name:
        queryset = queryset.filter(last_name=last_name)
    if cod_rol:
        queryset = queryset.filter(cod_rol=cod_rol)
    if id:
        queryset = queryset.filter(id=id)
    # Agrega aquí los demás filtros según los campos que deseas filtrar

    serializer = UserSerializer(queryset, many=True)

    # Obtener los datos serializados
    datos = serializer.data

    # Crear el libro de Excel
    wb = Workbook()
    ws = wb.active

    # Establecer el estilo del título
    titulo_font = Font(name='Arial', bold=True, size=16, color="FFFFFF")
    titulo_alignment = Alignment(horizontal='center', vertical='center')
    titulo_fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')

    # Combina las celdas del título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(datos[0]))

    # Escribe el título en la celda combinada
    titulo_cell = ws.cell(row=1, column=1, value='Lista de Usuarios')
    titulo_cell.font = titulo_font
    titulo_cell.alignment = titulo_alignment
    titulo_cell.fill = titulo_fill

    # Establecer el estilo de las celdas con borde
    borde = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Escribir los encabezados en el archivo de Excel
    encabezados = list(datos[0].keys())
    ws.append(encabezados)

    # Escribir los datos en el archivo de Excel
    for dato in datos:
        fila = []
        for clave, valor in dato.items():
            if isinstance(valor, list) and len(valor) == 0:
                valor = ""  # Reemplazar listas vacías por un valor vacío en Excel
            elif clave == 'cod_rol':
                try:
                    rol_id = valor
                    rol = Rol.objects.get(id=rol_id)
                    valor = rol.name_rol
                except Rol.DoesNotExist:
                    valor = ''  # Si el rol no existe, asignar un valor vacío
            elif clave == 'type_document':
                try:
                    type_document_id = valor
                    type_document = DocumentType.objects.get(id=type_document_id)
                    valor = type_document.name
                except DocumentType.DoesNotExist:
                    valor = ''  # Si el tipo de documento no existe, asignar un valor vacío
            elif clave == 'gender_user':
                try:
                    gender_id = valor
                    gender_n = gender.objects.get(id=gender_id)
                    valor = gender_n.name_gender
                except gender.DoesNotExist:
                    valor = ''  # Si el tipo de documento no existe, asignar un valor vacío
            fila.append(valor)
        ws.append(fila)

    # Aplicar estilo a las celdas con borde y colores
    for row in ws.iter_rows(min_row=2, min_col=1):  # Excluir la primera fila (encabezados)
        for cell in row:
            cell.border = borde
            cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # Ajustar el ancho de las columnas
    for columna in ws.columns:
        max_length = 0
        column = get_column_letter(columna[0].column)  # Obtener la letra de la columna
        for cell in columna:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Ajustar el ancho según el contenido
        ws.column_dimensions[column].width = adjusted_width

    # Crear la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="lista-usuarios.xlsx"'

    # Guardar el libro de Excel en el buffer
    buffer_excel = io.BytesIO()
    wb.save(buffer_excel)
    buffer_excel.seek(0)

    # Escribir el contenido del buffer en la respuesta HTTP
    response.write(buffer_excel.getvalue())

    return response