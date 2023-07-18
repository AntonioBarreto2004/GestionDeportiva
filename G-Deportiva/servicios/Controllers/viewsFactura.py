import io
from django.http import FileResponse
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from rest_framework.decorators import api_view
from rest_framework import status
from rest_framework.response import Response
from ..models import *
from django.shortcuts import get_object_or_404

def generar_factura(recibo_pago):
    servicio = recibo_pago.servicio

    nombre_servicio = servicio.nombreServicio
    descripcion_servicio = servicio.descripcionServicio
    valor_servicio = servicio.valorServicio
    fecha_pago = recibo_pago.fechaPago
    valor_pago = recibo_pago.valorPago

    total_pago = valor_servicio + valor_pago

    factura = {
        'nombre_servicio': nombre_servicio,
        'descripcion_servicio': descripcion_servicio,
        'valor_servicio': valor_servicio,
        'fecha_pago': fecha_pago,
        'valor_pago': valor_pago,
        'total_pago': total_pago
    }

    return factura


@api_view(['GET'])
def download_pdf_recibo_pago(request, pk):
    try:
        recibo_pago = get_object_or_404(ReciboPago, pk=pk)
        factura = generar_factura(recibo_pago)

        # Crear el PDF
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer)

        # Configurar el contenido del PDF
        p.setFont("Helvetica", 12)
        p.drawString(100, 700, "Factura de Pago")
        p.drawString(100, 650, "Nombre del Servicio: " + factura['nombre_servicio'])
        p.drawString(100, 600, "Descripción del Servicio: " + factura['descripcion_servicio'])
        p.drawString(100, 550, "Valor del Servicio: $" + str(factura['valor_servicio']))
        p.drawString(100, 500, "Fecha de Pago: " + str(factura['fecha_pago']))
        p.drawString(100, 450, "Valor del Pago: $" + str(factura['valor_pago']))
        p.drawString(100, 400, "Total a Pagar: $" + str(factura['total_pago']))

        p.showPage()
        p.save()

        # Volver al inicio del búfer y devolver la respuesta del archivo PDF
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='factura.pdf')
    except ReciboPago.DoesNotExist:
        return Response(
            data={
                'code': status.HTTP_404_NOT_FOUND,
                'message': 'Recibo de pago no encontrado.',
                'status': False
            },
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            data={
                'code': status.HTTP_500_INTERNAL_SERVER_ERROR,
                'message': 'Error al generar la factura: Ningún Recibo de Pago coincide con la consulta dada.',
                'status': False
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
def download_excel_recibo_pago(request, pk):
    try:
        recibo_pago = get_object_or_404(ReciboPago, pk=pk)
        factura = generar_factura(recibo_pago)

        # Crear el archivo Excel
        wb = Workbook()
        ws = wb.active

        # Configurar el contenido del archivo Excel
        ws['A1'] = "Factura de Pago"
        ws.merge_cells('A1:E1')
        ws['A3'] = "Nombre del Servicio"
        ws['B3'] = factura['nombre_servicio']
        ws['A4'] = "Descripción del Servicio"
        ws['B4'] = factura['descripcion_servicio']
        ws['A5'] = "Valor del Servicio"
        ws['B5'] = factura['valor_servicio']
        ws['A6'] = "Fecha de Pago"
        ws['B6'] = str(factura['fecha_pago'])
        ws['A7'] = "Valor del Pago"
        ws['B7'] = factura['valor_pago']
        ws['A8'] = "Total a Pagar"
        ws['B8'] = factura['total_pago']

        # Ajustar el ancho de las columnas
        column_letter = get_column_letter(2)
        ws.column_dimensions[column_letter].width = 20

        # Alinear el contenido al centro de las celdas
        align_center = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = align_center

        # Guardar el archivo Excel en memoria
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Devolver la respuesta del archivo Excel
        return FileResponse(excel_file, as_attachment=True, filename='factura.xlsx')
    except ReciboPago.DoesNotExist:
        return Response(
            data={
                'code': status.HTTP_404_NOT_FOUND,
                'message': 'Recibo de pago no encontrado.',
                'status': False
            },
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            data={
                'code': status.HTTP_500_INTERNAL_SERVER_ERROR,
                'message': 'Error al generar la factura: Ningún Recibo de Pago coincide con la consulta dada.',
                'status': False
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
